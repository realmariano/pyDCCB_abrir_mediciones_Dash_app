# Entradas de usuario
# --------------------------------------------------
archivo_de_medicion = '20210225_05_INTI_centro_200A60mV'
archivo_de_protocolo = 'protocolo_rev6'
# 'ENG' or 'ESP', this will impact on result because excel changes equations depending on language
archivo_de_protocolo_idioma = 'ENG'
n_estadistica_medicion = 20

print(" archivo de medicion = {} \n archivo de protocolo = {} \n idioma del excel = {} \n numero de mediciones para la estadistica = {}".format(archivo_de_medicion, archivo_de_protocolo, archivo_de_protocolo_idioma, n_estadistica_medicion))



# Carga de paquetes y definición de funciones
#------------------------------------------------------------------------------------------------
import pandas as pd
import io
import os
import sys
import traceback  # to indicate errors and exceptions 
import re
from inspect import currentframe, getframeinfo
import openpyxl as oxl
from openpyxl.formula.translate import Translator
print('package import finished')


def openFile(path2file):
    try:
        with open(path2File, mode='r') as f:  # para asegurar que se cierre correctamente uso with
            lines = f.readlines()
            return lines
    except Exception as exception:
        print("There was a problem with opening %s" % path2file)
        print('error in line {}'.format(
            inspect.getframeinfo(inspect.currentframe()).lineno))
        print(exception)
        raise

# Unused function, but perhaps we could use it in the next implementation
def block2df(dataBlock, col_names, num_rows=15):
    # returns block of data in data frame format, removes header and footer
    dfData = pd.read_csv(f, delimiter= ';', header=0, names=col_names, skiprows=83, engine='python', nrows=num_rows)
    return dfData

# The function reads and arrange the data 'list' of the file called by openFile
def read_mea_file(lines_of_mea_file):
    ln_preface = 81
    chunks = []
    qlist = []
    chunksInit =[]
    qlistInit = []
    chunksFinal = []
    qlistFinal = []
    nn = 0
    counter = 0
    ii = 0
    for line in lines:
        if (nn>ln_preface and line != '\n' and ii == 0):
            chunks.append(line)
            nn += 1
        elif (nn<ln_preface):
            chunksInit.append(line)
            nn +=1
        elif (nn>ln_preface and line == '\n'):
            ii = 1
            nn += 1
        elif (nn>ln_preface and ii == 1 and line != '***\n'):
            chunksFinal.append(line)
            nn += 1
        elif (line =='***\n'):
            ii = 0
            counter += 1
            nn = 0
            qlist.append(chunks)
            qlistInit.append(chunksInit)
            qlistFinal.append(chunksFinal)
            chunks = []
            chunksInit = []
            chunksFinal = []
        else:
            nn += 1
    return qlistInit, qlist, qlistFinal

def clearCells(workbook, cells):
    # example of how to input cells: for row in ws['A1:G37'] 
    for row in workbook[cells]:
        for cell in row:
            cell.value = None



## to be improved
""" ======================================================================

I removed the code below and kept a long protocol with lots of sheets, because it does not copy the charts, so its easiert to just fill and then remove the ones not used create as much sheets as needed
2 - 
~~~
for nn in range(len(qlist)):
  source = xl['(0)']
  target = xl.copy_worksheet(source)
~~~

3- En el archivo de protocolo hay que modificar la inclusión del nombre de archivo de mea, set y config

4- Incluir un método para que tome automáticamente si la medición se hace en aire o en el baño y corrija el valor de la resistencia (probablemente dos protocolos son necesarios para esto)

5- Modificar celda P10 del protocolo a: 
~~~
=IFS(NUMBERVALUE(C23)<30,30,NUMBERVALUE(C23)<100,300,NUMBERVALUE(C23)<300,3000) 
~~~

De forma tal que reconozca la relación del RE automáticamente.

6- El programa no está generando correctamente la salida
"""





# Entradas de usuario
# --------------------------------------------------
archivo_de_medicion = '20210225_05_INTI_centro_200A60mV'
archivo_de_protocolo = 'protocolo_rev6'
# 'ENG' or 'ESP', this will impact on result because excel changes equations depending on language
archivo_de_protocolo_idioma = 'ENG'
n_estadistica_medicion = 20

print(" archivo de medicion = {} \n archivo de protocolo = {} \n idioma del excel = {} \n numero de mediciones para la estadistica = {}".format(archivo_de_medicion, archivo_de_protocolo, archivo_de_protocolo_idioma, n_estadistica_medicion))


external_stylesheets = ['https://codepen.io/chriddyp/pen/bWLwgP.css']

app = dash.Dash(__name__, external_stylesheets=external_stylesheets)

app.layout = html.Div([
    dcc.Upload(
        id='upload-data',
        children=html.Div([
            'Drag and Drop o ',
            html.A('seleccione archivos')
        ]),
        style={
            'width': '50%',
            'height': '60px',
            'lineHeight': '60px',
            'borderWidth': '1px',
            'borderStyle': 'solid',
            'borderColor': '#6c6', 
            'backgroundColor': '#eee',
            'borderRadius': '5px',
            'textAlign': 'center',
            'margin': '20px'
        },
        # Allow multiple files to be uploaded
        multiple=True
    ),
    html.Div(id='output-data-upload'),
])


##########################################
#   Save data to xlsx file 
##########################################
# The protocolo_rev#_AAA.xlsx already has all the data and sheets necessary, to be filled

# select file to read
meaFile = archivo_de_medicion
path2File = meaFile + '.mea'
# open path2File
lines = openFile(path2File)
# read path2File and save meas to qlist, it has a different file in each list element qlist[i]
qlistInit, qlist, qlistFinal = read_mea_file(lines)

# Assign spreadsheet filename to `file`
xlsx_name = archivo_de_protocolo
language = archivo_de_protocolo_idioma # 'ENG' or 'ESP', this will impact on result because excel changes equations depending on language
xlsx_extension = '.xlsx'
xlsx_file = xlsx_name + '_' + language + xlsx_extension

# Load spreadsheet
xl = oxl.load_workbook(xlsx_file)

# sheet number sheetZero should be sheet named '(0)'
n_sheetZero = 4 # given the structure of the protocol.xlsx there are 4 sheets before (0)
sheet_name = xl.sheetnames[n_sheetZero]

# check if the sheets are in the right position, note that it depends on the order of the sheets that's why I include an if before the for, so it rises an exception if it is different from expected
if (sheet_name != '(0)'):
    cf = currentframe()
    raise Exception('File: {}, error in line{}: The xlsx shets have been moved, please check that you are using the original \' {} \' file'.format(getframeinfo(cf).filename, cf.f_lineno, xlsx_name))

# remove all extra sheets not needed
n_sheetnames = len(xl.sheetnames)
n_qlist = len(qlist)
for qq in reversed(range(n_sheetZero + n_qlist,n_sheetnames)):
    std= xl[xl.sheetnames[qq]]
    xl.remove(std)
    
# columns where the data of qlist should be saved
col1 = ['C','D','E','G','H','I','J','L','O','P']
# first row where the data of qlist should be saved
row1 = 83 
for qq in range(0,len(qlist)):
    # activate sheet, skips first n_sheetZero sheets and starts at sheet (0)
    sheet = xl[xl.sheetnames[n_sheetZero + qq]]

    # save data of qlist into the sheets
    ii = 0
    for dato in qlist[qq]:
        datosplit = dato.split(';')
        row2 = str(row1 + ii)
        sheet['B' + row2] = float(ii) # save meas number
        sheet[col1[0] + row2] = float(datosplit[0])
        sheet[col1[1] + row2] = float(datosplit[1])
        sheet[col1[2] + row2] = datosplit[2]
        # formulas to include
        sheet[col1[3] + row2] = "=$P$9*(1+$U$9*($R$15-20)+$V$9*($R$15-20)^2)"
        sheet[col1[4] + row2] = "="+ col1[0]+ str(row2)+ "*"+ col1[3]+ row2 
        sheet[col1[5] + row2] = "=("+col1[4]+str(row2)+"-$P$6*0,001)"
        sheet[col1[6] + row2] = "=("+col1[4]+str(row2)+"-$P$6*0.001)*1000000/($P$6*0.001)"

        if language == 'ENG':
            sheet[col1[7] + row2] = "=AVERAGE(" + col1[4] + str(row1) + ":" + col1[4] + row2 + ")"
            sheet[col1[8] + row2] = "=_xlfn.STDEV.P(" + col1[4] + str(row1) + ":" + col1[4] + row2 + ")"
            # its necessary to use the _xlfn only in the STDEV function, the rest will work fine. This will avoid excell incluiding an @ in front of the equations.
        elif language == 'ESP':
            sheet[col1[7] + row2] = "=PROMEDIO(" + col1[4] + str(row1) + ":" + col1[4] + row2 + ")"
            sheet[col1[8] + row2] = "=_xlfn.DESVEST.P(" + col1[4] + str(row1) + ":" + col1[4] + row2 + ")"
        else:
            try:
                raise ValueError()
            except:
                print(traceback.format_exc())
                print('Lenguaje seleccionado incorrecto o mal escrito, solo acepta ENG o ESP')
        ii +=1
    
    str_range =  col1[4] + str(len(qlist[qq]) -1 + row1 - n_estadistica_medicion) + ":" + col1[4] + str(len(qlist[qq]) -1 + row1) 
    if language == 'ENG':
        sheet[col1[8] + str(23)] = "=_xlfn.AVERAGE(" + str_range + ")" 
        sheet[col1[9] + str(23)] = "=_xlfn.STDEV.P(" + str_range + ")*1e6/O23" 
    elif language == 'ESP':
        sheet[col1[8] + str(23)] = "=_xlfn.PROMEDIO(" + str_range + ")"
        sheet[col1[9] + str(23)] = "=_xlfn.DESVEST.P(" + str_range + ")*1e6/O23" 
    
    # remove cells that are not in the measured range (they are preloaded in the xlsx protocol file)
    celdas_borrar_init = len(qlist[qq]) + row1
    celdas_borrar_final = 500 # I assume this is the last cell with values, if changed in the prtotocol this must be modify acordingly.
    celdas_str = 'A{initial}:O{final}'.format(initial=celdas_borrar_init, final=celdas_borrar_final)
    clearCells(sheet, celdas_str)
  
    # save data of qlistInit into the sheets
    jj = 1
    for datoInit in qlistInit[qq]:
        sheet[col1[0] + str(jj)] = datoInit.rstrip()
        jj += 1
    jj = 3
    for datoFinal in qlistFinal[qq]:
        sheet['F' + str(jj)] = datoFinal.rstrip()
        jj += 1

       
# Save data changes to new excel file
xl.save(meaFile + '_meas' + xlsx_extension)
print('The file has been saved as: {}'.format(os.getcwd() + '\\' + meaFile + '_meas' + xlsx_extension))



@app.callback(Output('output-data-upload', 'children'),
              Input('upload-data', 'contents'),
              State('upload-data', 'filename'),
              State('upload-data', 'last_modified'))
def update_output(list_of_contents, list_of_names, list_of_dates):
    if list_of_contents is not None:
        children = [
            parse_contents(c, n, d) for c, n, d in
            zip(list_of_contents, list_of_names, list_of_dates)]
        return children



if __name__ == '__main__':
    app.run_server(debug=True)
